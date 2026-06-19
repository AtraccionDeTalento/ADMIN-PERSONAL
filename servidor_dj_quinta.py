# ═══════════════════════════════════════════════════════════════════════════════
#  servidor_dj_quinta.py — Servidor Flask para Clasificador DJ Quinta Categoría
#  People Analytics USIL · Puerto :5010
#
#  Rutas:
#    GET  /                                    → UI principal
#    POST /api/v1/quinta/clasificar            → Upload de archivos PDF
#    POST /api/v1/quinta/clasificar-rutas      → Rutas locales (JSON)
#    POST /api/v1/quinta/clasificar-carpeta    → Carpeta local recursiva (JSON)
#    POST /api/v1/quinta/exportar-excel        → Genera .xlsx
#    POST /api/v1/f118/procesar-carpeta        → Extrae datos F-118 DTP
#    POST /api/v1/f118/exportar-excel          → Excel consolidado F-118
# ═══════════════════════════════════════════════════════════════════════════════

import os
import sys
from datetime import datetime
from pathlib import Path
from io import BytesIO

# Compatibilidad con pythonw.exe (sin consola) — stdout/stderr son None
if sys.stdout is None:
    sys.stdout = open(os.devnull, 'w', encoding='utf-8')
if sys.stderr is None:
    sys.stderr = open(os.devnull, 'w', encoding='utf-8')

if hasattr(sys.stdout, 'reconfigure'):
    try: sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception: pass
if hasattr(sys.stderr, 'reconfigure'):
    try: sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception: pass

try:
    from flask import Flask, render_template, request, jsonify, send_file
    from werkzeug.utils import secure_filename
    import openpyxl
    from clasificador_quinta import ClasificadorQuinta, _EXTENSIONES_SOPORTADAS, _EXTENSIONES_IMAGEN, _EXTENSIONES_PDF
    from procesador_ofertas import ProcesadorOfertas
    from procesador_combinado import ProcesadorCombinado
    from procesador_f118 import ProcesadorF118
except ImportError as e:
    print(f"\n[ERROR] Libreria faltante: {e}")
    print("Por favor ejecuta INICIAR.bat para instalar las dependencias.\n")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
app = Flask(__name__, template_folder=str(BASE_DIR / 'templates'))
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024   # 500 MB
app.config['UPLOAD_FOLDER'] = BASE_DIR / 'uploads_quinta'
app.config['UPLOAD_FOLDER'].mkdir(exist_ok=True)

clasificador = ClasificadorQuinta()

# ─────────────────────────────────────────────────────────────────────────────
# TAREAS EN SEGUNDO PLANO (DUAL)
# ─────────────────────────────────────────────────────────────────────────────
_DUAL_TASKS = {} # {id: {status, current, total, results, start_time}}

# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/')
@app.route('/quinta')
def index():
    return render_template('quinta_categoria.html')

# ─────────────────────────────────────────────────────────────────────────────
# API: UPLOAD (drag & drop / selector de archivos)
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/v1/quinta/clasificar', methods=['POST'])
def api_clasificar():
    """Recibe múltiples archivos subidos desde el navegador y los clasifica en paralelo."""
    if 'files' not in request.files:
        return jsonify({'status': 'ERROR', 'message': 'Sin archivos'}), 400

    files = request.files.getlist('files')
    rutas_guardadas = []

    for file in files:
        if not file.filename:
            continue
        ext = Path(file.filename).suffix.lower()
        if ext not in _EXTENSIONES_SOPORTADAS:
            continue
        filename = secure_filename(file.filename)
        filepath = app.config['UPLOAD_FOLDER'] / filename
        file.save(filepath)
        rutas_guardadas.append(filepath)

    if not rutas_guardadas:
        return jsonify({'status': 'ERROR', 'message': f'No se encontraron archivos válidos. Formatos soportados: {", ".join(_EXTENSIONES_SOPORTADAS)}'}), 400

    resultados = clasificador._procesar_paralelo(rutas_guardadas)
    resumen    = clasificador.generar_resumen(resultados)

    return jsonify({
        'status': 'OK',
        'resultados': resultados,
        'resumen': resumen
    }), 200

# ─────────────────────────────────────────────────────────────────────────────
# API: RUTAS LOCALES
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/v1/quinta/clasificar-rutas', methods=['POST'])
def api_clasificar_rutas():
    """
    Clasifica archivos por rutas locales.
    JSON: {"rutas": ["C:/ruta/archivo.pdf", "C:/carpeta/"]}
    """
    data = request.json
    if not data or 'rutas' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Se requiere campo "rutas"'}), 400

    rutas_raw = data['rutas']
    if not isinstance(rutas_raw, list) or len(rutas_raw) == 0:
        return jsonify({'status': 'ERROR', 'message': '"rutas" debe ser lista no vacía'}), 400

    archivos = []
    errores_previos = []

    for ruta_str in rutas_raw:
        ruta_str = ruta_str.strip().strip('"').strip("'")
        if not ruta_str:
            continue
        ruta_path = Path(ruta_str)
        if ruta_path.is_dir():
            encontrados = []
            for ext in _EXTENSIONES_SOPORTADAS:
                encontrados.extend(sorted(ruta_path.glob(f'*{ext}')))
                encontrados.extend(sorted(ruta_path.glob(f'*{ext.upper()}')))
            encontrados = list(dict.fromkeys(encontrados))
            if not encontrados:
                errores_previos.append(_err_result(ruta_str, f'Carpeta sin archivos soportados: {ruta_str}'))
            archivos.extend(encontrados)
        elif ruta_path.is_file() and ruta_path.suffix.lower() in _EXTENSIONES_SOPORTADAS:
            archivos.append(ruta_path)
        else:
            errores_previos.append(_err_result(ruta_str, f'Ruta no válida o formato no soportado: {ruta_str}'))

    resultados = clasificador._procesar_paralelo(archivos) + errores_previos
    resumen    = clasificador.generar_resumen(resultados)

    return jsonify({'status': 'OK', 'resultados': resultados, 'resumen': resumen}), 200

# ─────────────────────────────────────────────────────────────────────────────
# API: CARPETA RECURSIVA (método principal de la UI)
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/v1/quinta/clasificar-carpeta', methods=['POST'])
def api_clasificar_carpeta():
    """
    Escanea una carpeta raíz buscando declaraciones de quinta.
    Si hay subcarpetas por persona → analiza cada una y reporta también
    las personas sin declaración encontrada.
    JSON: {"carpeta": "C:/ruta/a/la/carpeta"}
    """
    data = request.json
    if not data or 'carpeta' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Se requiere campo "carpeta"'}), 400

    carpeta_str = data['carpeta'].strip().strip('"').strip("'")
    carpeta     = Path(carpeta_str)

    if not carpeta.exists():
        return jsonify({'status': 'ERROR', 'message': f'Carpeta no encontrada: {carpeta_str}'}), 404
    if not carpeta.is_dir():
        return jsonify({'status': 'ERROR', 'message': f'La ruta no es una carpeta: {carpeta_str}'}), 400

    # ── Detectar si hay subcarpetas por persona ───────────────────────────────
    subcarpetas = [d for d in sorted(carpeta.iterdir()) if d.is_dir()]

    if subcarpetas:
        # Modo persona: una carpeta por colaborador
        resultados = _procesar_por_personas(subcarpetas)
    else:
        # Modo plano: todos los archivos directo en la carpeta raíz
        archivos = []
        for ext in _EXTENSIONES_SOPORTADAS:
            archivos.extend(sorted(carpeta.glob(f'*{ext}')))
            archivos.extend(sorted(carpeta.glob(f'*{ext.upper()}')))
        archivos = list(dict.fromkeys(archivos))

        if not archivos:
            return jsonify({
                'status': 'OK', 'resultados': [],
                'resumen': clasificador.generar_resumen([]),
                'message': f'No se encontraron archivos soportados. Formatos: {", ".join(_EXTENSIONES_SOPORTADAS)}'
            }), 200
        resultados = clasificador._procesar_paralelo(archivos)
        resultados = [r for r in resultados if r['status'] in ('OK', 'WARNING')]

    resumen = clasificador.generar_resumen(resultados)
    return jsonify({
        'status': 'OK',
        'resultados': resultados,
        'resumen': resumen,
        'total_personas': len(subcarpetas) if subcarpetas else len(resultados)
    }), 200

# ─────────────────────────────────────────────────────────────────────────────
# API: EXPORTAR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/v1/quinta/exportar-excel', methods=['POST'])
def api_exportar_excel():
    """Genera un Excel .xlsx formateado con los resultados de clasificación."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.json
    if not data or 'resultados' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Sin datos para exportar'}), 400

    resultados = data['resultados']
    resumen    = data.get('resumen', {})

    COLOR_CAT = {
        '1A': 'FF2196F3',
        '1B': 'FF4CAF50',
        '2':  'FFFF9800',
        '3':  'FFF44336',
    }
    C_AZUL   = 'FF1A237E'
    C_AZUL2  = 'FF283593'
    C_BLANCO = 'FFFFFFFF'

    thin   = Side(style='thin', color='FFD0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── HOJA 1: DETALLE ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Clasificacion DJ Quinta'
    ws.sheet_view.showGridLines = False

    NUM_COLS = 6
    last_col = get_column_letter(NUM_COLS)  # 'F'

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'CLASIFICACION — DECLARACION JURADA QUINTA CATEGORIA'
    ws['A1'].font      = Font(bold=True, size=14, color=C_BLANCO)
    ws['A1'].fill      = PatternFill('solid', fgColor=C_AZUL)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = f'People Analytics USIL  ·  Generado: {datetime.now().strftime("%d/%m/%Y  %H:%M:%S")}'
    ws['A2'].font      = Font(size=10, color='FF888888', italic=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    # 6 columnas: N° | Persona | DNI (Declaración) | Nombre | Cat. | Descripción
    headers = [
        'N°', 'Persona (Carpeta)', 'DNI (Declaración)',
        'Nombre (Declaración)', 'Cat.', 'Descripción Categoría',
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(bold=True, size=11, color=C_BLANCO)
        c.fill      = PatternFill('solid', fgColor=C_AZUL2)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = border
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = 'A5'

    for idx, r in enumerate(resultados, 1):
        row      = 4 + idx
        cat      = r.get('categoria') or ''
        status   = r.get('status', 'ERROR')
        cat_lbl  = cat if status == 'OK' else ('?' if status == 'WARNING' else 'ERR')
        cat_desc = (r.get('categoria_info') or {}).get('nombre', r.get('mensaje', '-'))
        cat_clr  = COLOR_CAT.get(cat, 'FF9E9E9E')
        row_bg   = 'FFF5F7FF' if idx % 2 == 0 else C_BLANCO

        values = [
            idx,                                         # 1  N°
            r.get('persona') or r.get('nombre', '-'),   # 2  Persona (Carpeta)
            r.get('dni', '-'),                           # 3  DNI (Declaración)
            r.get('nombre', '-'),                        # 4  Nombre (Declaración)
            cat_lbl,                                     # 5  Cat.
            cat_desc,                                    # 6  Descripción Categoría
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border

            if col == 1:  # N°
                c.font      = Font(color='FF9E9E9E', size=10)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill      = PatternFill('solid', fgColor=row_bg)

            elif col == 2:  # Persona (Carpeta)
                c.font      = Font(bold=True, color=C_AZUL, size=11)
                c.alignment = Alignment(vertical='center', wrap_text=True)
                c.fill      = PatternFill('solid', fgColor=row_bg)

            elif col == 3:  # DNI (Declaración)
                c.font      = Font(color='FF333333', size=10)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill      = PatternFill('solid', fgColor=row_bg)

            elif col == 4:  # Nombre (Declaración)
                c.font      = Font(color='FF444444', size=10)
                c.alignment = Alignment(vertical='center', wrap_text=True)
                c.fill      = PatternFill('solid', fgColor=row_bg)

            elif col == 5:  # Cat. — color de categoría
                c.font      = Font(bold=True, color=C_BLANCO, size=12)
                c.fill      = PatternFill('solid', fgColor=cat_clr)
                c.alignment = Alignment(horizontal='center', vertical='center')

            else:  # col 6 — Descripción Categoría
                c.font      = Font(color='FF444444', size=10)
                c.alignment = Alignment(vertical='center', wrap_text=True)
                c.fill      = PatternFill('solid', fgColor=row_bg)

        ws.row_dimensions[row].height = 20

    # Anchos de columna: N° | Persona | DNI | Nombre | Cat. | Descripción
    col_widths = [5, 40, 16, 30, 8, 44]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f'A4:{last_col}{4 + len(resultados)}'

    # ── HOJA 2: RESUMEN ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Resumen')
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:C1')
    ws2['A1'] = 'RESUMEN ESTADISTICO'
    ws2['A1'].font      = Font(bold=True, size=13, color=C_BLANCO)
    ws2['A1'].fill      = PatternFill('solid', fgColor=C_AZUL)
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 32

    cat_counts = resumen.get('categorias') or {}
    res_data = [
        ('Total Procesados',                                               resumen.get('total_procesados', len(resultados)), 'FF1A237E'),
        ('Clasificados OK',                                                resumen.get('exitosos', 0),       'FF2E7D32'),
        ('Sin Clasificar',                                                 resumen.get('sin_clasificar', 0), 'FFE65100'),
        ('Errores',                                                        resumen.get('errores', 0),        'FFC62828'),
        None,
        ('Categoria 1A — Unico Empleador  (con renta quinta previa)',     cat_counts.get('1A', 0),          'FF2196F3'),
        ('Categoria 1B — Unico Empleador  (sin renta quinta previa)',     cat_counts.get('1B', 0),          'FF4CAF50'),
        ('Categoria 2  — Empleador Principal + otros ingresos',           cat_counts.get('2',  0),          'FFFF9800'),
        ('Categoria 3  — No es Empleador Principal',                      cat_counts.get('3',  0),          'FFF44336'),
    ]

    for i, item in enumerate(res_data, 3):
        if item is None:
            ws2.row_dimensions[i].height = 10
            continue
        label, val, color = item
        lc = ws2.cell(row=i, column=1, value=label)
        vc = ws2.cell(row=i, column=3, value=val)
        lc.font = Font(bold=True, color=color, size=11)
        vc.font = Font(bold=True, color=color, size=15)
        b   = Border(bottom=Side(style='thin', color='FFE8E8E8'))
        lc.border = b
        vc.border = b
        vc.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[i].height = 22

    ws2.column_dimensions['A'].width = 54
    ws2.column_dimensions['B'].width = 4
    ws2.column_dimensions['C'].width = 16

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"clasificacion_quinta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )

# ─────────────────────────────────────────────────────────────────────────────
# API: REINICIAR SERVIDOR
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/v1/quinta/reiniciar', methods=['POST'])
def api_reiniciar():
    """Reinicia el servidor Flask recargando el proceso Python."""
    import threading
    import subprocess

    def _reiniciar():
        import time
        time.sleep(0.5)
        # Ejecutar el mismo script de nuevo
        subprocess.Popen(
            [sys.executable, __file__],
            cwd=str(BASE_DIR),
            creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
        )
        # Terminar el proceso actual
        os._exit(0)

    threading.Thread(target=_reiniciar, daemon=True).start()
    return jsonify({'status': 'OK', 'message': 'Reiniciando...'}), 200

@app.route('/api/v1/utils/select-folder', methods=['POST'])
def api_select_folder():
    """Abre un diálogo nativo de Windows usando un proceso independiente para evitar bloqueos."""
    import subprocess
    try:
        # Usar el script independiente para evitar conflictos de hilos con Tkinter en Flask
        script_path = os.path.join(os.path.dirname(__file__), 'select_folder.py')
        res = subprocess.run([sys.executable, script_path], 
                             capture_output=True, text=True, check=False)
        
        folder = res.stdout.strip()
        if folder:
            return jsonify({'status': 'OK', 'path': folder}), 200
        return jsonify({'status': 'CANCEL', 'message': 'Selección cancelada'}), 200
    except Exception as e:
        return jsonify({'status': 'ERROR', 'message': str(e)}), 500

@app.route('/api/v1/utils/open-folder', methods=['POST'])
def api_open_folder():
    """Abre una carpeta local en el explorador de archivos."""
    data = request.json
    path = data.get('path', '').strip().strip('"').strip("'")
    if not path or not os.path.exists(path):
        return jsonify({'status': 'ERROR', 'message': 'Ruta no válida'}), 400
    
    try:
        if sys.platform == 'win32':
            os.startfile(path)
        elif sys.platform == 'darwin':
            import subprocess
            subprocess.Popen(['open', path])
        else:
            import subprocess
            subprocess.Popen(['xdg-open', path])
        return jsonify({'status': 'OK'}), 200
    except Exception as e:
        return jsonify({'status': 'ERROR', 'message': str(e)}), 500

# ─────────────────────────────────────────────────────────────────────────────
# API: CARTAS OFERTA (FUNCIONALIDAD 2)
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/v1/ofertas/procesar', methods=['POST'])
def api_procesar_ofertas():
    """
    Procesa cartas oferta desde una ruta origen a una destino.
    JSON: {"origen": "...", "destino": "..."}
    """
    data = request.json
    if not data or 'origen' not in data or 'destino' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Rutas origen y destino requeridas'}), 400
    
    origen = data['origen'].strip().strip('"').strip("'")
    destino = data['destino'].strip().strip('"').strip("'")
    
    procesador = ProcesadorOfertas(origen, destino)
    res = procesador.ejecutar()
    
    return jsonify(res), 200

@app.route('/api/v1/combinado/procesar', methods=['POST'])
def api_procesar_combinado():
    """
    Inicia el proceso dual en segundo plano.
    Retorna un task_id para seguimiento.
    """
    data = request.json
    if not data or 'origen' not in data or 'destino' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Rutas origen y destino requeridas'}), 400
    
    origen = data['origen'].strip().strip('"').strip("'")
    destino = data['destino'].strip().strip('"').strip("'")
    
    task_id = f"TASK-{datetime.now().strftime('%H%M%S')}-{os.urandom(2).hex().upper()}"
    _DUAL_TASKS[task_id] = {
        'status': 'RUNNING',
        'current': 0,
        'total': 0,
        'resultados': [],
        'start_time': datetime.now().isoformat(),
        'end_time': None,
        'message': 'Iniciando...'
    }

    def _run():
        import traceback
        try:
            def on_progress(curr, tot, res):
                _DUAL_TASKS[task_id]['current'] = curr
                _DUAL_TASKS[task_id]['total'] = tot
                _DUAL_TASKS[task_id]['resultados'].append(res)
                _DUAL_TASKS[task_id]['message'] = f"Procesando {curr}/{tot}"

            # Validar rutas antes de crear el procesador
            from pathlib import Path as _P
            if not _P(origen).exists():
                raise FileNotFoundError(f"Carpeta origen no encontrada: {origen}")
            subcarpetas_check = [d for d in _P(origen).iterdir() if d.is_dir()]
            if not subcarpetas_check:
                raise ValueError(f"La carpeta origen no tiene subcarpetas de personas: {origen}")

            procesador = ProcesadorCombinado(origen, destino)
            final_res = procesador.ejecutar(on_progress=on_progress)

            _DUAL_TASKS[task_id]['status'] = 'COMPLETED'
            _DUAL_TASKS[task_id]['end_time'] = datetime.now().isoformat()
            _DUAL_TASKS[task_id]['message'] = 'Completado con éxito'
        except Exception as e:
            _DUAL_TASKS[task_id]['status'] = 'ERROR'
            _DUAL_TASKS[task_id]['message'] = f"{type(e).__name__}: {e}"
            _DUAL_TASKS[task_id]['traceback'] = traceback.format_exc()

    import threading
    threading.Thread(target=_run, daemon=True).start()
    
    return jsonify({'status': 'OK', 'task_id': task_id}), 200

@app.route('/api/v1/combinado/status/<task_id>', methods=['GET'])
def api_get_combinado_status(task_id):
    """Retorna el estado y progreso de una tarea dual."""
    task = _DUAL_TASKS.get(task_id)
    if not task:
        return jsonify({'status': 'ERROR', 'message': 'Tarea no encontrada'}), 404
    
    return jsonify(task), 200

@app.route('/api/v1/ofertas/clasificar', methods=['POST'])
def api_clasificar_ofertas():
    """Procesa cartas oferta subidas directamente desde el navegador."""
    if 'files' not in request.files:
        return jsonify({'status': 'ERROR', 'message': 'Sin archivos'}), 400

    files = request.files.getlist('files')
    resultados = []
    
    # Carpeta temporal para procesamiento
    tmp_folder = BASE_DIR / 'tmp_ofertas'
    tmp_folder.mkdir(exist_ok=True)
    
    procesador = ProcesadorOfertas("", "") # Solo para usar sus métodos de triage

    for file in files:
        if not file.filename: continue
        filename = secure_filename(file.filename)
        filepath = tmp_folder / filename
        file.save(filepath)
        
        texto = procesador.extraer_texto(filepath)
        hallazgos = procesador.triage_contenido(texto)
        
        resultados.append({
            'persona': 'Carga Directa',
            'archivo_origen': filename,
            'hallazgos': hallazgos,
            'estado': 'PROCESADO_RAM',
            'timestamp': datetime.now().strftime("%H:%M:%S")
        })
        
        # Opcional: borrar el archivo después de procesar
        try: os.remove(filepath)
        except: pass

    return jsonify({'status': 'OK', 'resultados': resultados}), 200

@app.route('/api/v1/ofertas/procesar-rutas', methods=['POST'])
def api_procesar_ofertas_rutas():
    """Procesa una lista de rutas locales de cartas oferta."""
    data = request.json
    if not data or 'rutas' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Faltan rutas'}), 400

    rutas = data['rutas']
    resultados = []
    procesador = ProcesadorOfertas("", "")

    for r in rutas:
        path = Path(r.strip().strip('"').strip("'"))
        if not path.exists():
            resultados.append({'archivo_origen': path.name, 'estado': 'ERROR', 'mensaje': 'No existe'})
            continue
            
        texto = procesador.extraer_texto(path)
        hallazgos = procesador.triage_contenido(texto)
        
        resultados.append({
            'persona': 'Ruta Manual',
            'archivo_origen': path.name,
            'hallazgos': hallazgos,
            'ruta_destino': str(path.parent),
            'estado': 'PROCESADO_LOCAL',
            'timestamp': datetime.now().strftime("%H:%M:%S")
        })

    return jsonify({'status': 'OK', 'resultados': resultados}), 200

@app.route('/api/v1/ofertas/exportar-excel', methods=['POST'])
def api_exportar_excel_ofertas():
    """Genera un reporte Excel de las cartas oferta procesadas."""
    data = request.json
    if not data or 'resultados' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Sin datos'}), 400

    resultados = data['resultados']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Cartas Oferta"

    # Encabezados
    headers = ["UUID", "Archivo Origen", "Hallazgos Triage", "Persona/Ruta", "Estado", "Fecha/Hora"]
    ws.append(headers)

    # Estilos encabezado
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for res in resultados:
        ws.append([
            res.get('uuid', 'N/A'),
            res.get('archivo_origen', ''),
            ", ".join(res.get('hallazgos', [])),
            res.get('persona', 'Ruta Manual'),
            res.get('estado', ''),
            res.get('timestamp', '')
        ])

    output_path = BASE_DIR / "Reporte_Cartas_Oferta.xlsx"
    wb.save(output_path)
    
    return send_file(output_path, as_attachment=True)

# ─────────────────────────────────────────────────────────────────────────────
# API: F-118 DTP (FUNCIONALIDAD 3)
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/v1/f118/procesar-carpeta', methods=['POST'])
def api_f118_procesar_carpeta():
    """Escanea una carpeta raíz buscando archivos F_118_1_DTP.xlsx y extrae datos."""
    data = request.json
    if not data or 'carpeta' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Se requiere campo "carpeta"'}), 400

    carpeta_str = data['carpeta'].strip().strip('"').strip("'")
    carpeta = Path(carpeta_str)

    if not carpeta.exists():
        return jsonify({'status': 'ERROR', 'message': f'Carpeta no encontrada: {carpeta_str}'}), 404
    if not carpeta.is_dir():
        return jsonify({'status': 'ERROR', 'message': f'La ruta no es una carpeta: {carpeta_str}'}), 400

    procesador = ProcesadorF118()
    resultados = procesador.procesar_carpeta(carpeta_str)
    resumen = procesador.generar_resumen(resultados)

    return jsonify({
        'status': 'OK',
        'resultados': resultados,
        'resumen': resumen,
        'total_personas': len(resultados)
    }), 200


@app.route('/api/v1/f118/exportar-excel', methods=['POST'])
def api_f118_exportar_excel():
    """Genera un Excel consolidado con los datos extraídos de F-118."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.json
    if not data or 'resultados' not in data:
        return jsonify({'status': 'ERROR', 'message': 'Sin datos para exportar'}), 400

    resultados = data['resultados']
    resumen = data.get('resumen', {})

    C_AZUL = 'FF1A237E'
    C_AZUL2 = 'FF283593'
    C_BLANCO = 'FFFFFFFF'
    C_TEAL = 'FF00695C'
    thin = Side(style='thin', color='FFD0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── HOJA 1: DATOS CONSOLIDADOS ────────────────────────────────────────
    ws = wb.active
    ws.title = 'Datos Consolidados F-118'
    ws.sheet_view.showGridLines = False

    headers = [
        'N°', 'Persona (Carpeta)', 'Apellido Paterno', 'Apellido Materno',
        'Nombres', 'DNI', 'Correo', 'Celular', 'Distrito',
        'Cat. Quinta', 'Opción Quinta', 'Banco', 'N° Cuenta',
        'Correo DJ Tecno', 'Status', 'Archivo'
    ]
    NUM_COLS = len(headers)
    last_col = get_column_letter(NUM_COLS)

    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'EXTRACCIÓN F-118-1 DTP — DATOS DE PERSONAL'
    ws['A1'].font = Font(bold=True, size=14, color=C_BLANCO)
    ws['A1'].fill = PatternFill('solid', fgColor=C_TEAL)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = f'People Analytics USIL  ·  Generado: {datetime.now().strftime("%d/%m/%Y  %H:%M:%S")}'
    ws['A2'].font = Font(size=10, color='FF888888', italic=True)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = Font(bold=True, size=10, color=C_BLANCO)
        c.fill = PatternFill('solid', fgColor=C_AZUL2)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[4].height = 30
    ws.freeze_panes = 'A5'

    COLOR_CAT = {'1A': 'FF2196F3', '1B': 'FF4CAF50', '2': 'FFFF9800', '3': 'FFF44336', '?': 'FF9E9E9E'}

    for idx, r in enumerate(resultados, 1):
        row = 4 + idx
        res = r.get('resumen', {})
        fd = r.get('ficha_datos', {})
        ph = r.get('pago_haberes', {})
        dj = r.get('dj_tecnologias', {})
        q = r.get('ddjj_quinta', {})
        st = r.get('status', 'ERROR')
        row_bg = 'FFF5F7FF' if idx % 2 == 0 else C_BLANCO
        cat = q.get('categoria', '?') if st == 'OK' else '?'

        values = [
            idx,
            r.get('persona_carpeta', '-'),
            res.get('apellido_paterno', fd.get('apellido_paterno', '')),
            res.get('apellido_materno', fd.get('apellido_materno', '')),
            res.get('nombres', fd.get('nombres', '')),
            str(res.get('dni', fd.get('dni', q.get('dni', '')))),
            res.get('correo', fd.get('correo', '')),
            str(res.get('celular', fd.get('celular', ''))),
            res.get('distrito', fd.get('distrito', '')),
            cat,
            q.get('opcion_marcada', r.get('mensaje', '')),
            ph.get('banco', ''),
            ph.get('numero_cuenta', ''),
            dj.get('correo_personal', ''),
            st,
            r.get('archivo_origen', '-')
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            c.fill = PatternFill('solid', fgColor=row_bg)
            c.alignment = Alignment(vertical='center', wrap_text=True)
            c.font = Font(size=10)

            if col == 1:
                c.font = Font(color='FF9E9E9E', size=10)
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 2:
                c.font = Font(bold=True, color=C_AZUL, size=10)
            elif col == 10:
                cat_clr = COLOR_CAT.get(str(val), 'FF9E9E9E')
                c.font = Font(bold=True, color=C_BLANCO, size=11)
                c.fill = PatternFill('solid', fgColor=cat_clr)
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 15:
                if val == 'OK':
                    c.font = Font(bold=True, color='FF2E7D32', size=10)
                elif val == 'SIN_ARCHIVO':
                    c.font = Font(bold=True, color='FFE65100', size=10)
                else:
                    c.font = Font(bold=True, color='FFC62828', size=10)

        ws.row_dimensions[row].height = 20

    col_widths = [5, 38, 16, 16, 20, 12, 28, 14, 16, 8, 34, 14, 30, 28, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f'A4:{last_col}{4 + len(resultados)}'

    # ── HOJA 2: DDJJ QUINTA DETALLE ───────────────────────────────────────
    ws2 = wb.create_sheet('DDJJ Quinta Detalle')
    ws2.sheet_view.showGridLines = False

    h2 = ['N°', 'Persona', 'Apellidos y Nombres (DJ)', 'DNI', 'Domicilio',
          'Distrito', 'Op.1 Único', 'Op.1A No Percibió', 'Op.1B Sí Percibió',
          'Op.2 Principal', 'Op.3 No Principal', 'Categoría', 'Fecha Firma']

    ws2.merge_cells(f'A1:{get_column_letter(len(h2))}1')
    ws2['A1'] = 'DETALLE DDJJ QUINTA CATEGORÍA'
    ws2['A1'].font = Font(bold=True, size=13, color=C_BLANCO)
    ws2['A1'].fill = PatternFill('solid', fgColor=C_AZUL)
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 32

    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, size=10, color=C_BLANCO)
        c.fill = PatternFill('solid', fgColor=C_AZUL2)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws2.freeze_panes = 'A4'

    for idx, r in enumerate(resultados, 1):
        if r.get('status') != 'OK':
            continue
        row = ws2.max_row + 1
        q = r.get('ddjj_quinta', {})
        cat = q.get('categoria', '?')
        values = [
            idx, r.get('persona_carpeta', '-'),
            q.get('nombre_completo', ''), q.get('dni', ''),
            q.get('domicilio_calle', ''), q.get('domicilio_distrito', ''),
            'X' if q.get('opcion1_unico_empleador') else '',
            'X' if q.get('opcion1a_no_percibio') else '',
            'X' if q.get('opcion1b_si_percibio') else '',
            'X' if q.get('opcion2_empleador_principal') else '',
            'X' if q.get('opcion3_no_empleador_principal') else '',
            cat, q.get('fecha_firma', '')
        ]
        for col, val in enumerate(values, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col == 12:
                cat_clr = COLOR_CAT.get(str(val), 'FF9E9E9E')
                c.font = Font(bold=True, color=C_BLANCO, size=11)
                c.fill = PatternFill('solid', fgColor=cat_clr)

    for i, w in enumerate([5, 36, 30, 12, 20, 16, 10, 14, 14, 12, 14, 10, 16], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"extraccion_f118_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _procesar_por_personas(subcarpetas):
    """
    Para cada carpeta-persona: busca su declaración de quinta.
    Procesa en paralelo. Reporta 'Sin declaración' si no hay archivo quinta.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _mejor_resultado_persona(carpeta):
        persona = clasificador._extraer_persona_carpeta(carpeta / '_dummy')

        # Buscar todos los formatos soportados
        archivos = []
        for ext in _EXTENSIONES_SOPORTADAS:
            archivos.extend(sorted(carpeta.rglob(f'*{ext}')))
            archivos.extend(sorted(carpeta.rglob(f'*{ext.upper()}')))
        archivos = list(dict.fromkeys(archivos))

        # Priorizar archivos que parecen ser la declaración por su nombre
        def _prioridad_archivo(f):
            nombre = f.name.lower()
            if 'declaracion' in nombre or 'dj' in nombre or 'quinta' in nombre:
                if 'djtec' not in nombre: # Evitar djtec
                    return 0
            return 1

        archivos.sort(key=_prioridad_archivo)

        if not archivos:
            return {
                'status': 'ERROR', 'archivo': '', 'nombre': persona,
                'persona': persona,
                'categoria': None, 'categoria_info': None,
                'confianza': 0, 'metodo': 'sin_archivo',
                'mensaje': 'Sin archivo (PDF/imagen) en la carpeta'
            }

        # Identificar el DNI correcto de la persona (en paralelo con la búsqueda quinta)
        dni_doc = clasificador.identificar_dni_persona(carpeta, persona)

        # Intentar clasificar cada archivo; quedarse con el primero que sea quinta
        for archivo in archivos:
            r = clasificador.analizar_archivo(archivo)
            # Solo aceptar si el status es OK o si es un WARNING con confianza real
            if r['status'] == 'OK' or (r['status'] == 'WARNING' and r.get('confianza', 0) > 0):
                if not r.get('persona'):
                    r['persona'] = persona
                # Enriquecer con datos del DNI verificado
                if dni_doc:
                    r['dni_doc']        = dni_doc.get('numero')
                    r['dni_doc_nombre'] = dni_doc.get('nombre_completo')
                    r['dni_doc_archivo']= dni_doc.get('archivo_dni')
                    r['dni_doc_puntaje']= round(dni_doc.get('puntaje', 0), 2)
                    # El DNI del documento oficial (si existe) SIEMPRE tiene prioridad sobre lo escrito en la DJ Quinta
                    if dni_doc.get('numero'):
                        r['dni'] = dni_doc['numero']
                return r

        # Ningún archivo fue declaración de quinta
        resultado = {
            'status': 'ERROR', 'archivo': ', '.join(a.name for a in archivos[:3]),
            'nombre': persona, 'persona': persona,
            'categoria': None, 'categoria_info': None,
            'confianza': 0, 'metodo': 'no_es_quinta',
            'mensaje': 'No se encontró declaración de quinta en los archivos de la carpeta'
        }
        if dni_doc:
            resultado['dni_doc']        = dni_doc.get('numero')
            resultado['dni_doc_nombre'] = dni_doc.get('nombre_completo')
            resultado['dni_doc_archivo']= dni_doc.get('archivo_dni')
            resultado['dni_doc_puntaje']= round(dni_doc.get('puntaje', 0), 2)
        return resultado

    n       = len(subcarpetas)
    workers = min(8, n)
    resultados = [None] * n

    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(_mejor_resultado_persona, sc): i
                   for i, sc in enumerate(subcarpetas)}
        for f in as_completed(futures):
            resultados[futures[f]] = f.result()

    return resultados


def _err_result(archivo, mensaje):
    return {
        'status': 'ERROR', 'archivo': str(archivo),
        'nombre': '', 'persona': '',
        'categoria': None, 'categoria_info': None,
        'confianza': 0, 'metodo': 'error', 'mensaje': mensaje
    }

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print()
    print("=" * 65)
    print("  CLASIFICADOR DJ QUINTA CATEGORIA — People Analytics USIL")
    print("=" * 65)
    print(f"  URL:   http://localhost:5010")
    print(f"  APIs:")
    print(f"    POST /api/v1/quinta/clasificar           (upload archivos)")
    print(f"    POST /api/v1/quinta/clasificar-rutas     (rutas locales)")
    print(f"    POST /api/v1/quinta/clasificar-carpeta   (carpeta recursiva)")
    print(f"    POST /api/v1/quinta/exportar-excel       (genera .xlsx)")
    print("=" * 65)
    print("  Presiona Ctrl+C para detener.")
    print()

    app.run(host='127.0.0.1', port=5010, debug=False, use_reloader=False)
