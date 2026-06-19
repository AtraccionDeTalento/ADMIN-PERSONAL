# ═══════════════════════════════════════════════════════════════════════════════
#  procesador_f118.py — Extractor de datos del formulario F-118-1 (DTP)
#  People Analytics USIL
#
#  Lee archivos F_118_1_DTP.xlsx desde carpetas de colaboradores y extrae
#  datos de todas las hojas: Ficha de Datos, Pago de Haberes, Sistema de
#  Pensiones, DJ Tecnologías, Declaración Jurada, DDJJ 5TA, etc.
# ═══════════════════════════════════════════════════════════════════════════════

import os
import re
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Nombres de archivo que buscamos
_F118_PATTERNS = ['f_118', 'f118', 'f-118', 'ficha_datos', 'dtp']
_XLSX_EXT = {'.xlsx', '.xls'}


def _safe(val):
    """Convierte valor de celda a string limpio."""
    if val is None:
        return ''
    if isinstance(val, bool):
        return val
    s = str(val).strip()
    return s if s != '#VALUE!' else ''


def _bool_check(val):
    """Interpreta checkboxes: True/False o 'X'."""
    if isinstance(val, bool):
        return val
    if val is None:
        return False
    s = str(val).strip().upper()
    return s in ('TRUE', 'X', 'V', 'SI', 'SÍ', 'VERDADERO')


class ProcesadorF118:
    """Extrae datos estructurados de archivos F-118-1 DTP."""

    def buscar_archivos_f118(self, carpeta_raiz):
        """Busca archivos F-118 en subcarpetas de colaboradores."""
        carpeta = Path(carpeta_raiz)
        if not carpeta.exists() or not carpeta.is_dir():
            return []

        resultados = []
        subcarpetas = [d for d in sorted(carpeta.iterdir()) if d.is_dir()]

        if subcarpetas:
            for sub in subcarpetas:
                archivos = self._buscar_xlsx_en_carpeta(sub)
                if archivos:
                    for arch in archivos:
                        resultados.append({
                            'persona': sub.name,
                            'archivo': str(arch),
                            'nombre_archivo': arch.name
                        })
                else:
                    resultados.append({
                        'persona': sub.name,
                        'archivo': None,
                        'nombre_archivo': None
                    })
        else:
            # Archivos directos en la carpeta raíz
            for f in sorted(carpeta.glob('*.xlsx')):
                if self._es_f118(f):
                    resultados.append({
                        'persona': carpeta.name,
                        'archivo': str(f),
                        'nombre_archivo': f.name
                    })
        return resultados

    def _buscar_xlsx_en_carpeta(self, carpeta):
        """Busca archivos xlsx F-118 dentro de una carpeta."""
        encontrados = []
        for f in sorted(carpeta.rglob('*.xlsx')):
            if self._es_f118(f):
                encontrados.append(f)
        return encontrados

    def _es_f118(self, archivo):
        """Determina si un archivo es un F-118 por su nombre."""
        if archivo.name.startswith('~$'):
            return False
        name = archivo.stem.lower().replace(' ', '_').replace('-', '_')
        for pat in _F118_PATTERNS:
            if pat in name:
                return True
        return False

    def procesar_carpeta(self, carpeta_raiz, on_progress=None):
        """Procesa todos los F-118 encontrados en la carpeta."""
        archivos_info = self.buscar_archivos_f118(carpeta_raiz)
        total = len(archivos_info)
        resultados = []

        for idx, info in enumerate(archivos_info, 1):
            if info['archivo']:
                try:
                    datos = self.extraer_datos(info['archivo'])
                    datos['persona_carpeta'] = info['persona']
                    datos['status'] = 'OK'
                    datos['archivo_origen'] = info['nombre_archivo']
                except Exception as e:
                    datos = {
                        'persona_carpeta': info['persona'],
                        'status': 'ERROR',
                        'mensaje': str(e),
                        'archivo_origen': info['nombre_archivo']
                    }
            else:
                datos = {
                    'persona_carpeta': info['persona'],
                    'status': 'SIN_ARCHIVO',
                    'mensaje': 'No se encontró archivo F-118 en la carpeta',
                    'archivo_origen': None
                }

            resultados.append(datos)
            if on_progress:
                on_progress(idx, total, datos)

        return resultados

    def extraer_datos(self, archivo_path):
        """Extrae todos los datos de un archivo F-118-1 DTP."""
        if not openpyxl:
            raise ImportError("openpyxl no está instalado")

        wb = openpyxl.load_workbook(str(archivo_path), data_only=True)
        hojas_encontradas = wb.sheetnames

        datos = {
            'hojas_detectadas': hojas_encontradas,
            'ficha_datos': {},
            'pago_haberes': {},
            'sistema_pensiones': {},
            'dj_tecnologias': {},
            'declaracion_jurada': {},
            'ddjj_quinta': {},
        }

        # Extraer de cada hoja
        for nombre_hoja in hojas_encontradas:
            ws = wb[nombre_hoja]
            nl = nombre_hoja.lower().strip()

            if 'ficha de datos' in nl:
                datos['ficha_datos'] = self._extraer_ficha_datos(ws)
            elif 'pago de haberes' in nl or 'pago' in nl:
                datos['pago_haberes'] = self._extraer_pago_haberes(ws)
            elif 'pensiones' in nl:
                datos['sistema_pensiones'] = self._extraer_pensiones(ws)
            elif 'tecnolog' in nl:
                datos['dj_tecnologias'] = self._extraer_dj_tecnologias(ws)
            elif nl.startswith('declaraci') and 'jurada' in nl:
                datos['declaracion_jurada'] = self._extraer_declaracion_jurada(ws)
            elif 'ddjj' in nl or '5ta' in nl or 'quinta' in nl:
                datos['ddjj_quinta'] = self._extraer_ddjj_quinta(ws)

        wb.close()

        # Datos consolidados
        fd = datos['ficha_datos']
        dj5 = datos['ddjj_quinta']
        datos['resumen'] = {
            'apellido_paterno': fd.get('apellido_paterno', ''),
            'apellido_materno': fd.get('apellido_materno', ''),
            'nombres': fd.get('nombres', ''),
            'nombre_completo': f"{fd.get('apellido_paterno', '')} {fd.get('apellido_materno', '')} {fd.get('nombres', '')}".strip(),
            'dni': fd.get('dni', dj5.get('dni', '')),
            'correo': fd.get('correo', ''),
            'celular': fd.get('celular', ''),
            'distrito': fd.get('distrito', ''),
            'categoria_quinta': dj5.get('categoria', ''),
            'opcion_quinta': dj5.get('opcion_marcada', ''),
        }

        return datos

    def _extraer_ficha_datos(self, ws):
        """Extrae datos de la hoja Ficha de Datos."""
        d = {}
        try:
            d['apellido_paterno'] = _safe(ws['A5'].value)
            d['apellido_materno'] = _safe(ws['D5'].value)
            d['nombres'] = _safe(ws['G5'].value)

            # Fecha nacimiento
            dia = _safe(ws['A8'].value)
            mes = _safe(ws['B8'].value)
            anio = _safe(ws['C8'].value)
            d['fecha_nacimiento'] = f"{dia}/{mes}/{anio}" if dia else ''

            d['pais_nacimiento'] = _safe(ws['D8'].value)
            d['depto_nacimiento'] = _safe(ws['F8'].value)
            d['provincia_nacimiento'] = _safe(ws['H8'].value)
            d['distrito_nacimiento'] = _safe(ws['I8'].value)
            d['estado_civil'] = _safe(ws['L7'].value)

            # DNI
            c10 = _safe(ws['C10'].value)
            d['dni'] = str(c10).strip() if c10 else ''
            d['correo'] = _safe(ws['D10'].value)
            d['celular'] = _safe(ws['J10'].value)

            # Dirección
            d['tipo_via'] = _safe(ws['B18'].value) if ws['B18'].value else ''
            d['nombre_via'] = _safe(ws['B19'].value)
            d['numero'] = _safe(ws['F19'].value)
            d['departamento'] = _safe(ws['H18'].value)
            d['provincia'] = _safe(ws['H19'].value)
            d['distrito'] = _safe(ws['H21'].value)
            d['interior'] = _safe(ws['F21'].value)

            # Info familiar
            familia = []
            for r in range(26, 31):
                parentesco = _safe(ws.cell(row=r, column=1).value)
                nombre_fam = _safe(ws.cell(row=r, column=3).value)
                if parentesco and nombre_fam:
                    familia.append({
                        'parentesco': parentesco,
                        'nombre': nombre_fam,
                        'ocupacion': _safe(ws.cell(row=r, column=7).value),
                    })
            d['familia'] = familia

            # Contacto emergencia
            d['emergencia_nombre'] = _safe(ws['A32'].value).replace('NOMBRE: ', '') if ws['A32'].value else ''
            d['emergencia_parentesco'] = _safe(ws['E32'].value).replace('PARENTESCO: ', '') if ws['E32'].value else ''
            d['emergencia_telefono'] = _safe(ws['I32'].value).replace('TELÉFONO: ', '').replace('TEL\u00c9FONO: ', '') if ws['I32'].value else ''

        except Exception as e:
            d['_error'] = str(e)
        return d

    def _extraer_pago_haberes(self, ws):
        """Extrae datos de Pago de Haberes."""
        d = {}
        try:
            # Banco seleccionado
            bcp = _bool_check(ws['D5'].value)
            scotia = _bool_check(ws['F5'].value) if ws['F5'].value else False
            bbva = _bool_check(ws['H5'].value) if ws['H5'].value else False

            if bcp or str(_safe(ws['D5'].value)).upper() == 'X':
                d['banco'] = 'BCP'
            elif scotia or str(_safe(ws['F5'].value)).upper() == 'X':
                d['banco'] = 'SCOTIABANK'
            elif bbva or str(_safe(ws['H5'].value)).upper() == 'X':
                d['banco'] = 'BBVA'
            else:
                d['banco'] = _safe(ws['C6'].value) or 'No especificado'

            d['numero_cuenta'] = _safe(ws['E7'].value)
            d['numero_cci'] = _safe(ws['E8'].value) if ws['E8'].value else ''
        except Exception as e:
            d['_error'] = str(e)
        return d

    def _extraer_pensiones(self, ws):
        """Extrae datos de Sistema de Pensiones."""
        d = {}
        try:
            # Detectar qué opción está marcada
            d['opcion'] = ''
            d['detalle'] = ''

            # Opción 1: Jubilado
            # Opción 2: Afiliado AFP
            for r in range(16, 18):
                for c in [3, 5, 7, 9]:
                    val = ws.cell(row=r+1, column=c).value
                    if val and str(val).strip().upper() == 'X':
                        afp_name = _safe(ws.cell(row=r+1, column=c-1).value) or _safe(ws.cell(row=r, column=c).value)
                        d['opcion'] = 'AFP'
                        # Identify which AFP
                        col_letter = ws.cell(row=r+1, column=c).column_letter
                        # Check row 17 for AFP names
                        if c == 3 or c == 4:
                            d['detalle'] = 'HABITAT'
                        elif c == 5 or c == 6:
                            d['detalle'] = 'INTEGRA'
                        elif c == 7 or c == 8:
                            d['detalle'] = 'PRIMA'
                        elif c == 9 or c == 10:
                            d['detalle'] = 'PROFUTURO'

            # Check checkboxes for AFP row 17
            for c_idx, nombre in [(3, 'HABITAT'), (5, 'INTEGRA'), (7, 'PRIMA'), (9, 'PROFUTURO')]:
                v17 = ws.cell(row=17, column=c_idx+1).value
                if v17 and str(v17).strip().upper() == 'X':
                    d['opcion'] = 'AFP'
                    d['detalle'] = nombre

            # Check ONP options
            v21 = _safe(ws['B21'].value)
            if v21 and ('4' in str(v21) or 'ONP' in str(v21).upper()):
                pass  # Check if marked

        except Exception as e:
            d['_error'] = str(e)
        return d

    def _extraer_dj_tecnologias(self, ws):
        """Extrae datos de DJ para uso de Tecnologías."""
        d = {}
        try:
            d['correo_personal'] = _safe(ws['D11'].value)
            d['correo_alternativo'] = _safe(ws['D12'].value)
            d['dni'] = _safe(ws['C18'].value)

            # Fecha
            dia = _safe(ws['C14'].value)
            mes = _safe(ws['D14'].value)
            anio = _safe(ws['E14'].value)
            d['fecha'] = f"{dia} {mes} {anio}".strip()
        except Exception as e:
            d['_error'] = str(e)
        return d

    def _extraer_declaracion_jurada(self, ws):
        """Extrae datos de la Declaración Jurada general."""
        d = {}
        try:
            d['correo_personal'] = _safe(ws['C32'].value)
            d['correo_institucional'] = _safe(ws['C33'].value)
        except Exception as e:
            d['_error'] = str(e)
        return d

    def _extraer_ddjj_quinta(self, ws):
        """Extrae datos de la DDJJ 5TA (Quinta Categoría)."""
        d = {}
        try:
            # Datos del colaborador
            d['apellido1'] = _safe(ws['E10'].value)
            d['apellido2'] = _safe(ws['G10'].value)
            d['nombres'] = _safe(ws['I10'].value)
            d['nombre_completo'] = f"{d['apellido1']} {d['apellido2']} {d['nombres']}".strip()

            d['domicilio_calle'] = _safe(ws['E11'].value)
            d['domicilio_numero'] = _safe(ws['F11'].value)
            d['domicilio_depto'] = _safe(ws['G11'].value)
            d['domicilio_distrito'] = _safe(ws['H11'].value)
            d['dni'] = _safe(ws['E12'].value)

            # Opciones de quinta categoría (búsqueda en rango para tolerar cambios de formato)
            def _marcado_en_rango(filas, columnas):
                for r in filas:
                    for c in columnas:
                        if _bool_check(ws[f"{c}{r}"].value):
                            return True
                return False

            op1 = _marcado_en_rango([16, 17, 18], ['B', 'C', 'D'])
            op1a = _marcado_en_rango([19, 20], ['B', 'C', 'D', 'E'])
            op1b = _marcado_en_rango([21, 22], ['B', 'C', 'D', 'E'])
            op2 = _marcado_en_rango([24, 25, 26, 27], ['B', 'C', 'D'])
            op3 = _marcado_en_rango([28, 29, 30, 31], ['B', 'C', 'D'])

            d['opcion1_unico_empleador'] = op1
            d['opcion1a_no_percibio'] = op1a
            d['opcion1b_si_percibio'] = op1b
            d['opcion2_empleador_principal'] = op2
            d['opcion3_no_empleador_principal'] = op3

            # Clasificar categoría
            if op1 and op1a:
                d['categoria'] = '1A'
                d['opcion_marcada'] = 'Único empleador - NO percibió renta quinta'
            elif op1 and op1b:
                d['categoria'] = '1B'
                d['opcion_marcada'] = 'Único empleador - SÍ percibió renta quinta'
            elif op1 and not op1a and not op1b:
                d['categoria'] = 'Requiere observación'
                d['opcion_marcada'] = 'Único empleador - Sin sub-opción marcada'
            elif op2:
                d['categoria'] = '2'
                d['opcion_marcada'] = 'Empleador principal + otros ingresos'
            elif op3:
                d['categoria'] = '3'
                d['opcion_marcada'] = 'No es empleador principal'
            else:
                d['categoria'] = 'Requiere observación'
                d['opcion_marcada'] = 'Sin opción marcada'

            # Fecha firma
            dia = _safe(ws['D40'].value)
            mes = _safe(ws['F40'].value)
            anio = _safe(ws['H40'].value)
            d['fecha_firma'] = f"{dia} {mes} {anio}".strip() if dia else ''

        except Exception as e:
            d['_error'] = str(e)
        return d

    def generar_resumen(self, resultados):
        """Genera estadísticas resumidas."""
        total = len(resultados)
        ok = sum(1 for r in resultados if r.get('status') == 'OK')
        sin_archivo = sum(1 for r in resultados if r.get('status') == 'SIN_ARCHIVO')
        errores = sum(1 for r in resultados if r.get('status') == 'ERROR')

        cats = {}
        for r in resultados:
            if r.get('status') == 'OK':
                cat = r.get('ddjj_quinta', {}).get('categoria', 'Requiere observación')
                cats[cat] = cats.get(cat, 0) + 1

        return {
            'total': total,
            'exitosos': ok,
            'sin_archivo': sin_archivo,
            'errores': errores,
            'categorias': cats,
        }
